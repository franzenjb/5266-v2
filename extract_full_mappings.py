#!/usr/bin/env python3
import openpyxl
import json

# Open the Statistics Excel file
wb = openpyxl.load_workbook('/Users/jefffranzen/5266-v2/planning-docs/Disaster5266CollectionTool-Statistics.xlsx', data_only=True)

# Extract the CountyStateRegions_21 sheet
sheet = wb['CountyStateRegions_21']

# Build the mappings
region_county_chapter_map = {}
all_regions = set()
all_chapters = set()

# Skip header row, start from row 2
for row in range(2, sheet.max_row + 1):
    chapter = sheet.cell(row=row, column=1).value
    region = sheet.cell(row=row, column=2).value
    county_state = sheet.cell(row=row, column=3).value
    
    if chapter and region and county_state:
        # Clean up the values
        chapter = str(chapter).strip()
        region = str(region).strip()
        county_state = str(county_state).strip()
        
        # Add to sets for unique values
        all_regions.add(region)
        all_chapters.add(chapter)
        
        # Build the mapping structure
        if region not in region_county_chapter_map:
            region_county_chapter_map[region] = {
                'counties': [],
                'chapters': set()
            }
        
        region_county_chapter_map[region]['counties'].append({
            'name': county_state,
            'chapter': chapter
        })
        region_county_chapter_map[region]['chapters'].add(chapter)

# Convert sets to lists for JSON serialization
for region in region_county_chapter_map:
    region_county_chapter_map[region]['chapters'] = sorted(list(region_county_chapter_map[region]['chapters']))

# Create the output structure
output = {
    'regions': sorted(list(all_regions)),
    'total_regions': len(all_regions),
    'total_chapters': len(all_chapters),
    'region_details': region_county_chapter_map
}

# Save to JSON file
with open('/Users/jefffranzen/5266-v2/region_county_chapter_data.json', 'w') as f:
    json.dump(output, f, indent=2)

# Print summary
print(f"Extracted {len(all_regions)} regions")
print(f"Found {len(all_chapters)} unique chapters")
print(f"\nFirst 5 regions with county counts:")
for i, region in enumerate(sorted(all_regions)[:5]):
    county_count = len(region_county_chapter_map[region]['counties'])
    chapter_count = len(region_county_chapter_map[region]['chapters'])
    print(f"  {region}: {county_count} counties, {chapter_count} chapters")

print("\nData saved to region_county_chapter_data.json")

# Also extract just Florida region for our demo
florida_regions = [r for r in all_regions if 'Florida' in r]
if florida_regions:
    print(f"\nFlorida regions found: {florida_regions}")
    for fl_region in florida_regions:
        print(f"  {fl_region}: {len(region_county_chapter_map[fl_region]['counties'])} counties")