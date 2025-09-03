#!/usr/bin/env python3
import openpyxl
import json
import sys

# Open the Statistics Excel file
wb = openpyxl.load_workbook('/Users/jefffranzen/5266-v2/planning-docs/Disaster5266CollectionTool-Statistics.xlsx', data_only=True)

# Print all sheet names first
print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
print()

# Look for county/chapter mappings in common sheets
def extract_data_from_sheet(sheet_name):
    if sheet_name not in wb.sheetnames:
        return None
    
    sheet = wb[sheet_name]
    data = []
    
    # Read first 100 rows and 20 columns to find patterns
    for row in range(1, min(sheet.max_row + 1, 100)):
        row_data = []
        for col in range(1, min(sheet.max_column + 1, 20)):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        # Only include non-empty rows
        if any(row_data):
            data.append(row_data)
    
    return data

# Check key sheets that might contain the mappings
sheets_to_check = [
    'Start Here Statistics',
    'Counties',
    'Chapters', 
    'Regions',
    'County-Chapter',
    'Master',
    'Setup',
    'Data',
    'Lists',
    'Validation'
]

print("Searching for county/chapter/region data...\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Look for sheets with county or chapter data
    if any(keyword in sheet_name.lower() for keyword in ['count', 'chapter', 'region', 'start', 'list', 'data', 'validation', 'setup']):
        print(f"\n=== Sheet: {sheet_name} ===")
        
        # Sample first 20 rows to understand structure
        for row in range(1, min(21, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    row_data.append(str(cell_value)[:50])  # Truncate long values
            
            if row_data:
                print(f"Row {row}: {row_data}")