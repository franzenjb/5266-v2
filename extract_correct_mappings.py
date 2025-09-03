#!/usr/bin/env python3
import openpyxl
import json

# Open the Statistics Excel file
wb = openpyxl.load_workbook('/Users/jefffranzen/5266-v2/planning-docs/Disaster5266CollectionTool-Statistics.xlsx', data_only=True)

# Extract the CountyStateRegions_21 sheet
sheet = wb['CountyStateRegions_21']

# Build the mappings
region_county_chapter_map = {}
all_regions = set()
all_chapters = set()
county_to_chapter = {}

# Skip header row, start from row 2
for row in range(2, min(sheet.max_row + 1, 5000)):  # Limit to prevent too much data
    chapter = sheet.cell(row=row, column=2).value  # Column B
    region = sheet.cell(row=row, column=3).value   # Column C
    county_state = sheet.cell(row=row, column=4).value  # Column D
    
    if chapter and region and county_state:
        # Clean up the values
        chapter = str(chapter).strip()
        region = str(region).strip()
        county_state = str(county_state).strip()
        
        # Add to sets for unique values
        all_regions.add(region)
        all_chapters.add(chapter)
        
        # Store county to chapter mapping
        county_to_chapter[county_state] = chapter
        
        # Build the mapping structure
        if region not in region_county_chapter_map:
            region_county_chapter_map[region] = {
                'counties': [],
                'chapters': set()
            }
        
        region_county_chapter_map[region]['counties'].append({
            'name': county_state,
            'chapter': chapter
        })
        region_county_chapter_map[region]['chapters'].add(chapter)

# Convert sets to lists for JSON serialization
for region in region_county_chapter_map:
    region_county_chapter_map[region]['chapters'] = sorted(list(region_county_chapter_map[region]['chapters']))
    # Sort counties by name
    region_county_chapter_map[region]['counties'].sort(key=lambda x: x['name'])

# Create the output structure
output = {
    'regions': sorted(list(all_regions)),
    'total_regions': len(all_regions),
    'total_chapters': len(all_chapters),
    'total_counties': len(county_to_chapter),
    'region_details': region_county_chapter_map,
    'county_to_chapter': county_to_chapter
}

# Save to JSON file
with open('/Users/jefffranzen/5266-v2/region_county_chapter_data.json', 'w') as f:
    json.dump(output, f, indent=2)

# Print summary
print(f"Extracted {len(all_regions)} regions")
print(f"Found {len(all_chapters)} unique chapters")
print(f"Found {len(county_to_chapter)} county-chapter mappings")

print(f"\nFirst 10 regions with county counts:")
for i, region in enumerate(sorted(all_regions)[:10]):
    county_count = len(region_county_chapter_map[region]['counties'])
    chapter_count = len(region_county_chapter_map[region]['chapters'])
    print(f"  {i+1}. {region}: {county_count} counties, {chapter_count} chapters")

# Look for Florida-specific regions
florida_regions = [r for r in all_regions if 'Florida' in r]
if florida_regions:
    print(f"\n=== Florida regions found ===")
    for fl_region in florida_regions:
        counties = region_county_chapter_map[fl_region]['counties']
        print(f"\n{fl_region}: {len(counties)} counties")
        print("Sample counties:")
        for county in counties[:5]:
            print(f"  - {county['name']} -> {county['chapter']}")