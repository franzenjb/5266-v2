#!/usr/bin/env python3
import openpyxl
import json

# Open the Statistics Excel file  
wb = openpyxl.load_workbook('/Users/jefffranzen/5266-v2/planning-docs/Disaster5266CollectionTool-Statistics.xlsx', data_only=True)

# Extract the CountyStateRegions_21 sheet
sheet = wb['CountyStateRegions_21']

# Build the mappings
regions_data = {}
all_counties_chapters = {}

# Skip header row, start from row 2
for row in range(2, min(sheet.max_row + 1, 5000)):
    chapter = sheet.cell(row=row, column=2).value  
    region = sheet.cell(row=row, column=3).value   
    county_state = sheet.cell(row=row, column=4).value 
    
    if chapter and region and county_state:
        chapter = str(chapter).strip()
        region = str(region).strip()
        county_state = str(county_state).strip()
        
        # Build region structure
        if region not in regions_data:
            regions_data[region] = {
                'counties': {},
                'chapters': set()
            }
        
        # Extract just county name without state for display
        county_display = county_state.split(',')[0].strip()
        
        regions_data[region]['counties'][county_state] = chapter
        regions_data[region]['chapters'].add(chapter)
        all_counties_chapters[county_state] = chapter

# Convert to JavaScript format
js_output = "// Region, County, and Chapter Database\n"
js_output += "// Extracted from Red Cross Form 5266 Excel file\n\n"
js_output += "const redCrossRegionsData = {\n"

for region in sorted(regions_data.keys()):
    js_output += f'  "{region}": {{\n'
    js_output += f'    countyCount: {len(regions_data[region]["counties"])},\n'
    js_output += f'    chapters: {json.dumps(sorted(list(regions_data[region]["chapters"])))},\n'
    js_output += '    counties: {\n'
    
    # Add first 200 counties per region to keep size manageable
    for i, (county, chapter) in enumerate(sorted(regions_data[region]['counties'].items())[:200]):
        js_output += f'      "{county}": "{chapter}"'
        if i < len(regions_data[region]['counties']) - 1 and i < 199:
            js_output += ','
        js_output += '\n'
    
    js_output += '    }\n'
    js_output += '  },\n'

js_output = js_output.rstrip(',\n') + '\n};\n\n'

# Add a list of all regions
js_output += "const allRegions = [\n"
for region in sorted(regions_data.keys()):
    js_output += f'  "{region}",\n'
js_output = js_output.rstrip(',\n') + '\n];\n'

# Save to file
with open('/Users/jefffranzen/5266-v2/redcross_data.js', 'w') as f:
    f.write(js_output)

print(f"Generated JavaScript data file with {len(regions_data)} regions")
print(f"Total counties across all regions: {len(all_counties_chapters)}")

# Show Florida summary
florida_regions = [r for r in regions_data.keys() if 'Florida' in r]
print(f"\nFlorida regions in database:")
for region in florida_regions:
    print(f"  - {region}: {len(regions_data[region]['counties'])} counties, {len(regions_data[region]['chapters'])} chapters")